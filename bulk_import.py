"""
Bulk import of farms and flights from a CSV or Excel sheet.

Two jobs, mirroring parsing.py's approach to the DroneDeploy export:

  1. Read a .csv, .xlsx or .xlsm upload into plain rows, tolerating however the
     person happened to label the columns — case, spacing, punctuation and the
     common synonyms are all normalised, so "Farm Name", "farm_name" and "FARM"
     all land on the same field.
  2. Validate each row on its own and say what would happen to it — create,
     update, or an error naming the cell at fault. One bad row never stops the
     rest of the sheet.

Matching follows the rules the app already enforces elsewhere: a farm is
identified by its name, and a flight by farm + season + flight number, which is
the same triple the flight form describes as unique. A row that matches an
existing record updates it; a row that matches nothing creates one.

A blank cell on an update means "leave this as it is", not "clear it". A part-
filled sheet is the normal way somebody corrects two phone numbers across forty
farms, and reading those blanks as deletions would quietly empty the record.
"""
import csv
import io
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except Exception:                      # pragma: no cover - optional dependency
    XLSX_AVAILABLE = False

CSV_EXTS = {".csv"}
XLSX_EXTS = {".xlsx", ".xlsm"}
ALLOWED_EXTS = CSV_EXTS | XLSX_EXTS

# A sheet this size is already far past what anyone maintains by hand, and the
# preview round-trips the parsed rows through the browser, so the cap keeps that
# payload sane as much as it guards the database.
MAX_ROWS = 500


# ---------------------------------------------------------------- columns
def _norm(header):
    """Fold a column heading down to bare lowercase letters and digits."""
    return "".join(ch for ch in str(header or "").lower() if ch.isalnum())


# Every accepted spelling of each field, normalised. Longest match wins simply
# because each key is exact — there is no prefix matching to get wrong.
FARM_COLUMNS = {
    "name": ["name", "farm", "farmname", "farmtitle"],
    "crop": ["crop", "cropname", "croptype"],
    "acreage": ["acreage", "acres", "size", "area", "fieldsize"],
    "location": ["location", "place", "region", "address", "county"],
    "farmer_name": ["farmername", "farmer", "contactname", "contact", "owner"],
    "farmer_email": ["farmeremail", "email", "contactemail", "emailaddress"],
    "farmer_phone": ["farmerphone", "phone", "whatsapp", "contactphone",
                     "phonenumber", "mobile", "whatsappphone"],
    "dronedeploy_project_url": ["dronedeployprojecturl", "dronedeploy",
                                "dronedeployurl", "dronedeploylink",
                                "projecturl", "maplink"],
    "notes": ["notes", "note", "comment", "comments", "remarks"],
}

FLIGHT_COLUMNS = {
    "farm": ["farm", "farmname", "name"],
    "season": ["season", "seasonyear", "seasoncode"],
    "flight_number": ["flightnumber", "flightno", "flight", "flightnum", "no"],
    "flights_planned": ["flightsplanned", "planned", "plannedflights",
                        "totalflights", "flightsinseason"],
    "crop": ["crop", "cropname", "croptype"],
    "acreage": ["acreage", "acres", "size", "area", "fieldsize"],
    "flight_date": ["flightdate", "date", "flown", "flownon", "dateflown"],
    "status": ["status", "stage", "state"],
    "dronedeploy_project_url": ["dronedeployprojecturl", "dronedeploy",
                                "dronedeployurl", "dronedeploylink",
                                "projecturl", "maplink"],
}

# What the downloadable template offers, in the order a person fills them in.
FARM_TEMPLATE = ["name", "crop", "acreage", "location", "farmer_name",
                 "farmer_email", "farmer_phone", "dronedeploy_project_url", "notes"]
FLIGHT_TEMPLATE = ["farm", "season", "flight_number", "flights_planned", "crop",
                   "acreage", "flight_date", "status", "dronedeploy_project_url"]

TEMPLATE_HEADINGS = {
    "name": "Farm name", "crop": "Crop", "acreage": "Acreage",
    "location": "Location", "farmer_name": "Farmer name",
    "farmer_email": "Farmer email", "farmer_phone": "Farmer phone",
    "dronedeploy_project_url": "DroneDeploy project URL", "notes": "Notes",
    "farm": "Farm name", "season": "Season", "flight_number": "Flight number",
    "flights_planned": "Flights planned", "flight_date": "Flight date",
    "status": "Status",
}

TEMPLATE_SAMPLE = {
    "farms": ["Kilimo Bora Farm", "Maize", "42", "Naromoru, Kenya", "John Mwangi",
              "john@example.com", "+254712345678",
              "https://www.dronedeploy.com/app2/sites/example", ""],
    "flights": ["Kilimo Bora Farm", "2026LR", "1", "12", "Maize", "42",
                "2026-03-14", "Draft", ""],
}

FLIGHT_STATUSES = ["Draft", "Ready for Review", "Approved", "Sent"]


def _column_map(headers, spec):
    """Map each column index in the sheet to the field it feeds, if any."""
    lookup = {}
    for field, aliases in spec.items():
        for a in aliases:
            lookup.setdefault(a, field)
    mapping, seen = {}, set()
    for i, h in enumerate(headers):
        field = lookup.get(_norm(h))
        # First column wins when a sheet repeats a heading, so a stray duplicate
        # column cannot silently override what was already read.
        if field and field not in seen:
            mapping[i] = field
            seen.add(field)
    return mapping


# ---------------------------------------------------------------- reading
def read_rows(file_bytes, filename):
    """
    Return (headers, rows) from a CSV or Excel upload, as lists of strings.

    Raises ValueError with a message meant for the person who uploaded it.
    """
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in XLSX_EXTS:
        if not XLSX_AVAILABLE:
            raise ValueError("Excel files need the openpyxl package on the server. "
                             "Save the sheet as CSV and upload that instead.")
        return _read_xlsx(file_bytes)
    if ext in CSV_EXTS:
        return _read_csv(file_bytes)
    raise ValueError(f"Unsupported file type: {ext or 'unknown'}. "
                     "Upload a .csv, .xlsx or .xlsm file.")


def _read_csv(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    # Sheets exported on a comma-decimal locale come out semicolon-separated, so
    # the delimiter is sniffed rather than assumed.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [[(c or "").strip() for c in r] for r in reader]
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        raise ValueError("That file has no rows in it.")
    return rows[0], rows[1:]


def _read_xlsx(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(values_only=True):
        row = [_cell(c) for c in raw]
        if any(c for c in row):
            rows.append(row)
    wb.close()
    if not rows:
        raise ValueError("That sheet has no rows in it.")
    return rows[0], rows[1:]


def _cell(value):
    """Excel cell to string, without Excel's decimal and date decorations."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))          # 42.0 typed as 42 reads back as 42
    return str(value).strip()


# ---------------------------------------------------------------- values
def clean_float(raw):
    """Acreage as typed: '42', '42.5', '42,5', '38 ac' all read as a number."""
    s = str(raw or "").strip().replace(",", ".")
    s = "".join(ch for ch in s if ch.isdigit() or ch in ".-")
    if not s or s in (".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_int(raw):
    s = str(raw or "").strip()
    s = "".join(ch for ch in s if ch.isdigit() or ch == "-")
    if not s or s == "-":
        return None
    try:
        return int(s)
    except ValueError:
        return None


DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
                "%d.%m.%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y")


def clean_date(raw):
    s = str(raw or "").strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clean_status(raw):
    """Snap a status onto the four the app recognises, however it was typed."""
    s = str(raw or "").strip()
    if not s:
        return None
    for valid in FLIGHT_STATUSES:
        if _norm(valid) == _norm(s):
            return valid
    return ""                          # present but unrecognised — caller errors


# ---------------------------------------------------------------- planning
def plan_farms(headers, rows, existing_by_name):
    """
    Work out what each row would do, without touching the database.

    `existing_by_name` maps a lowercased farm name to its id, so this module
    stays free of queries in the same way report_data does.
    """
    mapping = _column_map(headers, FARM_COLUMNS)
    plans = []
    if "name" not in mapping.values():
        return plans, ("No farm name column found. The sheet needs a column "
                       "headed 'Farm name' (or 'name'). Download the template "
                       "below to see the exact headings.")

    seen_in_file = {}
    for n, raw in enumerate(rows, start=2):        # row 1 is the heading row
        values = {field: (raw[i].strip() if i < len(raw) and raw[i] else "")
                  for i, field in mapping.items()}
        name = values.get("name", "").strip()
        plan = {"row": n, "fields": values, "label": name or "(no name)",
                "action": None, "error": None, "target_id": None}

        if not name:
            plan["error"] = "No farm name in this row."
        elif name.lower() in seen_in_file:
            plan["error"] = (f"The same farm appears earlier in this file, on "
                             f"row {seen_in_file[name.lower()]}.")
        else:
            seen_in_file[name.lower()] = n
            if values.get("acreage") and clean_float(values["acreage"]) is None:
                plan["error"] = f"Acreage '{values['acreage']}' is not a number."
            else:
                existing = existing_by_name.get(name.lower())
                plan["action"] = "update" if existing else "create"
                plan["target_id"] = existing

        if plan["error"]:
            plan["action"] = "error"
        plans.append(plan)
    return plans, None


def plan_flights(headers, rows, farms_by_name, flight_keys):
    """
    As plan_farms, for flights.

    `farms_by_name` maps a lowercased farm name to its id; `flight_keys` maps
    (farm_id, season_lower, flight_number) to a flight id — the same triple the
    flight form calls unique.
    """
    mapping = _column_map(headers, FLIGHT_COLUMNS)
    plans = []
    have = set(mapping.values())
    missing = [c for c in ("farm", "season", "flight_number") if c not in have]
    if missing:
        pretty = ", ".join(TEMPLATE_HEADINGS[m] for m in missing)
        return plans, (f"The sheet is missing a required column: {pretty}. "
                       "Download the template below to see the exact headings.")

    seen_in_file = {}
    for n, raw in enumerate(rows, start=2):
        values = {field: (raw[i].strip() if i < len(raw) and raw[i] else "")
                  for i, field in mapping.items()}
        farm_name = values.get("farm", "").strip()
        season = values.get("season", "").strip()
        number = clean_int(values.get("flight_number"))
        label = (f"{farm_name or '(no farm)'} · {season or '(no season)'} · "
                 f"flight {number if number is not None else '?'}")
        plan = {"row": n, "fields": values, "label": label,
                "action": None, "error": None, "target_id": None, "farm_id": None}

        farm_id = farms_by_name.get(farm_name.lower()) if farm_name else None
        key = (farm_id, season.lower(), number)

        if not farm_name:
            plan["error"] = "No farm name in this row."
        elif farm_id is None:
            plan["error"] = (f"No farm called '{farm_name}' exists yet. "
                             "Import the farms first, or correct the spelling.")
        elif not season:
            plan["error"] = "No season in this row (e.g. 2026LR)."
        elif number is None:
            plan["error"] = (f"Flight number '{values.get('flight_number', '')}' "
                             "is not a whole number.")
        elif number < 1:
            plan["error"] = "Flight number must be 1 or more."
        elif key in seen_in_file:
            plan["error"] = (f"The same farm, season and flight number appears "
                             f"earlier in this file, on row {seen_in_file[key]}.")
        elif values.get("acreage") and clean_float(values["acreage"]) is None:
            plan["error"] = f"Acreage '{values['acreage']}' is not a number."
        elif values.get("flights_planned") and clean_int(values["flights_planned"]) is None:
            plan["error"] = (f"Flights planned '{values['flights_planned']}' "
                             "is not a whole number.")
        elif values.get("flight_date") and clean_date(values["flight_date"]) is None:
            plan["error"] = (f"Date '{values['flight_date']}' is not one this "
                             "reads. Use 2026-03-14 or 14/03/2026.")
        elif values.get("status") and clean_status(values["status"]) == "":
            plan["error"] = (f"Status '{values['status']}' is not one of "
                             + ", ".join(FLIGHT_STATUSES) + ".")
        else:
            seen_in_file[key] = n
            existing = flight_keys.get(key)
            plan["action"] = "update" if existing else "create"
            plan["target_id"] = existing
            plan["farm_id"] = farm_id

        if plan["error"]:
            plan["action"] = "error"
        plans.append(plan)
    return plans, None


def summarise(plans):
    return {
        "to_add": sum(1 for p in plans if p["action"] == "create"),
        "to_update": sum(1 for p in plans if p["action"] == "update"),
        "to_skip": sum(1 for p in plans if p["action"] == "error"),
        "total": len(plans),
    }


# ---------------------------------------------------------------- templates
def template_csv(kind):
    """A starter sheet with the exact headings the importer reads."""
    fields = FARM_TEMPLATE if kind == "farms" else FLIGHT_TEMPLATE
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([TEMPLATE_HEADINGS[f] for f in fields])
    w.writerow(TEMPLATE_SAMPLE[kind])
    return out.getvalue()
